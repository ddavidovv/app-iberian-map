"""
Generate Zones Master file from Excel and SVG sources

This script:
1. Reads zones from Excel (customIbericMap_v28.xlsm - "zones" sheet)
2. Reads zone IDs from SVG (public/assets/iberian_map.svg)
3. Compares and creates a master zones file with status indicators
4. Outputs to public/data/zones-master.json
"""

import openpyxl
import json
import re
from typing import Dict, List, Set

# Known code mappings (SVG code -> Excel code)
CODE_MAPPINGS = {
    'ES-AC': 'ES-C',   # A Coruña
    'ES-ME': 'ES-ML',  # Melilla
    'ES-PA': 'ES-P',   # Palencia
}

REVERSE_MAPPINGS = {v: k for k, v in CODE_MAPPINGS.items()}


def load_excel_zones(excel_path: str) -> Dict[str, dict]:
    """Load zones from Excel file"""
    print("Loading Excel zones...")

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    if 'zones' not in wb.sheetnames:
        raise ValueError(f"Sheet 'zones' not found in Excel. Available: {wb.sheetnames}")

    ws = wb['zones']

    # Get headers
    headers = [cell.value for cell in ws[1]]

    # Load all zones
    zones = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # If first cell has value
            zone_data = {}
            for idx, header in enumerate(headers):
                if header and idx < len(row):
                    zone_data[header] = row[idx]

            zone_code = zone_data.get('zone_code')
            if zone_code:
                zones[zone_code] = zone_data

    print(f"  Loaded {len(zones)} zones from Excel")

    # Filter ES and PT zones
    es_pt_zones = {k: v for k, v in zones.items()
                   if v.get('country_code') in ['ES', 'PT']}

    print(f"  ES/PT zones: {len(es_pt_zones)}")

    return es_pt_zones


def load_svg_zones(svg_path: str) -> Dict[str, dict]:
    """Load zone IDs and names from SVG"""
    print("Loading SVG zones...")

    with open(svg_path, 'r', encoding='utf-8') as f:
        svg_content = f.read()

    # Find all elements with IDs
    pattern = r'id="([^"]+)"(?:[^>]*?)data-name="([^"]*)"'
    matches = re.findall(pattern, svg_content, re.DOTALL)

    svg_zones = {}
    for zone_id, zone_name in matches:
        # Filter out background elements
        if not zone_id.startswith('cuadrado_') and (zone_id.startswith('ES-') or zone_id.startswith('PT-')):
            svg_zones[zone_id] = {
                'svg_id': zone_id,
                'svg_name': zone_name
            }

    print(f"  Loaded {len(svg_zones)} ES/PT zones from SVG")

    return svg_zones


def generate_zones_master(excel_zones: Dict, svg_zones: Dict) -> Dict:
    """Generate master zones file with status"""
    print("\nGenerating zones master...")

    master = {}

    # Process all SVG zones
    for svg_id, svg_data in svg_zones.items():
        # Check if there's a direct match in Excel
        excel_data = excel_zones.get(svg_id)

        # Check if there's a mapped code
        mapped_code = CODE_MAPPINGS.get(svg_id)
        if not excel_data and mapped_code:
            excel_data = excel_zones.get(mapped_code)

        if excel_data:
            # Zone exists in both
            status = 'ok'
            warning_reason = None
            excel_equivalent = None

            # Check if it's a mapped code (inconsistent)
            if mapped_code:
                status = 'warning'
                warning_reason = f"Código inconsistente con Excel ({mapped_code})"
                excel_equivalent = mapped_code

            master[svg_id] = {
                'zone_code': svg_id,
                'zone_name': excel_data.get('zone_name', '').upper(),
                'country_code': excel_data.get('country_code', ''),
                'zone_ccaa': excel_data.get('zone_ccaa', ''),
                'zone_group': excel_data.get('zone_group', ''),
                'zone_expreg': excel_data.get('zone_expreg', ''),
                'in_svg': True,
                'svg_id': svg_data['svg_id'],
                'svg_name': svg_data['svg_name'],
                'status': status,
                'warning_reason': warning_reason,
                'excel_equivalent': excel_equivalent
            }
        else:
            # Zone only in SVG
            master[svg_id] = {
                'zone_code': svg_id,
                'zone_name': svg_data['svg_name'].upper(),
                'country_code': svg_id.split('-')[0],
                'zone_ccaa': '',
                'zone_group': '',
                'zone_expreg': '',
                'in_svg': True,
                'svg_id': svg_data['svg_id'],
                'svg_name': svg_data['svg_name'],
                'status': 'svg_only',
                'warning_reason': 'Zona sin datos en Excel'
            }

    # Process Excel zones not in SVG (only if not already covered by reverse mapping)
    for excel_code, excel_data in excel_zones.items():
        # Skip if already processed via reverse mapping
        if excel_code in REVERSE_MAPPINGS:
            continue

        if excel_code not in master:
            master[excel_code] = {
                'zone_code': excel_code,
                'zone_name': excel_data.get('zone_name', '').upper(),
                'country_code': excel_data.get('country_code', ''),
                'zone_ccaa': excel_data.get('zone_ccaa', ''),
                'zone_group': excel_data.get('zone_group', ''),
                'zone_expreg': excel_data.get('zone_expreg', ''),
                'in_svg': False,
                'status': 'excel_only',
                'warning_reason': 'Zona no presente en SVG (no renderizable)'
            }

    # Count by status
    status_counts = {}
    for zone in master.values():
        status = zone['status']
        status_counts[status] = status_counts.get(status, 0) + 1

    print(f"\n  Total zones in master: {len(master)}")
    print(f"  Status breakdown:")
    for status, count in sorted(status_counts.items()):
        print(f"    - {status}: {count}")

    return master


def main():
    print("=" * 80)
    print("ZONES MASTER GENERATOR")
    print("=" * 80)

    # Load data
    excel_zones = load_excel_zones('customIbericMap_v28.xlsm')
    svg_zones = load_svg_zones('public/assets/iberian_map.svg')

    # Generate master
    zones_master = generate_zones_master(excel_zones, svg_zones)

    # Save to file
    output_path = 'public/data/zones-master.json'
    print(f"\nSaving to {output_path}...")

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(zones_master, f, indent=2, ensure_ascii=False)

    print("OK Done!")

    # Show warnings
    warnings = [z for z in zones_master.values() if z['status'] in ['warning', 'svg_only']]
    if warnings:
        print(f"\n⚠ {len(warnings)} zones with warnings:")
        for zone in warnings:
            print(f"  - {zone['zone_code']}: {zone['warning_reason']}")


if __name__ == '__main__':
    main()
