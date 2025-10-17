import openpyxl
import json

# Load the Excel file
wb = openpyxl.load_workbook('customIbericMap_v28.xlsm', data_only=True)

# Check if 'zones' sheet exists
if 'zones' in wb.sheetnames:
    ws = wb['zones']

    # Get headers (first row)
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    print("Headers found:")
    print(headers)
    print("\n" + "="*80 + "\n")

    # Get all data
    zones_data = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0]:  # If first cell has value
            row_data = {}
            for idx, header in enumerate(headers):
                if header and idx < len(row):
                    row_data[header] = row[idx]
            zones_data.append(row_data)

    print(f"Total zones found: {len(zones_data)}")
    print("\n" + "="*80 + "\n")

    # Show first 5 zones as sample
    print("Sample zones (first 5):")
    for zone in zones_data[:5]:
        print(json.dumps(zone, indent=2, ensure_ascii=False))
        print("-" * 40)

    # Save all zones to JSON
    with open('scripts/zones_from_excel.json', 'w', encoding='utf-8') as f:
        json.dump(zones_data, f, indent=2, ensure_ascii=False)

    print(f"\nAll zones saved to scripts/zones_from_excel.json")

    # Get unique zone IDs
    zone_ids = [z.get('id') or z.get('zone_id') or z.get('ID') for z in zones_data if z]
    zone_ids = [z for z in zone_ids if z]

    print(f"\nTotal unique zone IDs: {len(zone_ids)}")
    print("Zone IDs sample:", zone_ids[:10])

else:
    print("Sheet 'zones' not found!")
    print("Available sheets:", wb.sheetnames)
