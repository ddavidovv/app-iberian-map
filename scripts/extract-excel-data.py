"""
Script para extraer datos del Excel de Portfolio Ibérico y generar map-config.json

Este script lee el archivo Excel con las hojas de origen (Península, Canarias, Baleares, Islas Portugal)
y genera la configuración JSON para el mapa interactivo.
"""

import openpyxl
import json
import sys

# Configuración de encoding
sys.stdout.reconfigure(encoding='utf-8')

# Mapeo de categorías de destino a códigos de zona
# Este mapeo necesitará ajustes basados en el negocio
ZONE_MAPPINGS = {
    'peninsula': {
        'Provincial': [],  # Se calculará dinámicamente según la zona de origen
        'Regional': [],  # Requiere definición de regiones
        'Peninsular': [
            'ES-A', 'ES-AB', 'ES-AL', 'ES-AV', 'ES-B', 'ES-BA', 'ES-BI', 'ES-BU',
            'ES-C', 'ES-CA', 'ES-CC', 'ES-CO', 'ES-CR', 'ES-CS', 'ES-CU', 'ES-GI',
            'ES-GR', 'ES-GU', 'ES-H', 'ES-HU', 'ES-J', 'ES-L', 'ES-LE', 'ES-LO',
            'ES-LU', 'ES-M', 'ES-MA', 'ES-MU', 'ES-NA', 'ES-O', 'ES-OR', 'ES-P',
            'ES-PO', 'ES-S', 'ES-SA', 'ES-SE', 'ES-SG', 'ES-SO', 'ES-SS', 'ES-T',
            'ES-TE', 'ES-TO', 'ES-V', 'ES-VA', 'ES-VI', 'ES-Z', 'ES-ZA'
        ],
        'Peninsular Plus': [],  # Similar a Peninsular, requiere definición específica
        'Portugal': ['PT-AV', 'PT-BE', 'PT-BG', 'PT-BR', 'PT-CB', 'PT-CO', 'PT-EV',
                     'PT-FA', 'PT-GU', 'PT-LE', 'PT-LI', 'PT-PA', 'PT-PO', 'PT-SA',
                     'PT-SE', 'PT-VC', 'PT-VI', 'PT-VR'],
        'Isla Mayor': ['ES-GC-LA', 'ES-TF-TF', 'ES-PM-MA', 'ES-PM-IB'],  # Canarias y Baleares mayores
        'Isla Menor': ['ES-GC-FU', 'ES-GC-LZ', 'ES-TF-GO', 'ES-TF-HI', 'ES-TF-LP',
                       'ES-PM-ME', 'ES-PM-FO'],  # Islas menores
        'InterIslas': [],  # Tráfico entre islas
        'Madeira': ['PT-MA-MA', 'PT-MA-PS'],
        'Azores': ['PT-AZ-CO', 'PT-AZ-FA', 'PT-AZ-FL', 'PT-AZ-GR', 'PT-AZ-PI',
                   'PT-AZ-SA', 'PT-AZ-SJ', 'PT-AZ-SM', 'PT-AZ-TE'],
        'Ceuta / Melilla / Gibraltar / Andorra': ['ES-CE', 'ES-ML'],
    },
    'canarias': {
        'Origen\nIsla Mayor': ['ES-GC-LA', 'ES-TF-TF'],
        'Origen\nIsla Menor': ['ES-GC-FU', 'ES-GC-LZ', 'ES-TF-GO', 'ES-TF-HI', 'ES-TF-LP'],
        'InterIslas Mayores': [],
        'InterIslas \nResto de Islas': [],
        'Misma Isla Mayor': [],
    },
    'baleares': {
        'Origen\nIsla Mayor': ['ES-PM-MA', 'ES-PM-IB'],
        'Origen\nIsla Menor': ['ES-PM-ME', 'ES-PM-FO'],
        'InterIslas': [],
        'Misma Isla': [],
    },
    'islas_portugal': {
        'Azores-Madeira / Madeira - Azores': []
    }
}

# Mapeo de códigos de baremo
BAREMO_CODES = {
    'Provincial': 'PRO',
    'Regional': 'REG',
    'Peninsular': 'PEN',
    'Peninsular Plus': 'PEL',
    'Portugal': 'PTC',
    'Isla Mayor': 'CAM',  # Canarias/Baleares Mayor
    'Isla Menor': 'BAM',  # Baleares/Canarias Menor
    'InterIslas': 'PEN',
    'Madeira': 'PTI',
    'Azores': 'PTI',
    'Ceuta / Melilla / Gibraltar / Andorra': 'CEU',
    'Misma Isla': 'PRO',
    'Misma Isla Mayor': 'PRO',
}


def extract_products_from_sheet(ws, origin_type):
    """
    Extrae la configuración de productos de una hoja de origen

    Args:
        ws: Worksheet de openpyxl
        origin_type: Tipo de origen ('peninsula', 'canarias', 'baleares', 'islas_portugal')

    Returns:
        dict: Diccionario con productos y sus configuraciones
    """
    products_config = {}

    # Obtener las cabeceras de columnas (fila 6 para destinos)
    header_row = 6
    headers = []
    for cell in ws[header_row]:
        if cell.value:
            headers.append((cell.column, str(cell.value).strip()))

    print(f"  Cabeceras encontradas: {len(headers)}")

    # Leer productos (empiezan en fila 11 aproximadamente)
    for row_idx in range(11, ws.max_row + 1):
        row = ws[row_idx]

        # Columna F (índice 6) tiene el nombre del producto
        product_name_cell = row[5]  # Índice 5 = columna F
        if not product_name_cell.value:
            continue

        product_name = str(product_name_cell.value).strip()

        # Filtrar filas que no son productos
        if not product_name or product_name.startswith('servicios'):
            continue

        # Inicializar producto si no existe
        if product_name not in products_config:
            products_config[product_name] = {
                'destinations': {}
            }

        # Leer cada columna de destino
        for col_idx, header in headers:
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value

            # Si hay una marca (☑, ✓, etc.) o valor positivo, el servicio está disponible
            is_available = value and str(value).strip() in ['☑', '✓', 'x', 'X']

            if is_available:
                if header not in products_config[product_name]['destinations']:
                    products_config[product_name]['destinations'][header] = True

    return products_config


def generate_map_config(excel_path, output_path):
    """
    Genera el archivo map-config.json desde el Excel

    Args:
        excel_path: Ruta al archivo Excel
        output_path: Ruta donde guardar el JSON generado
    """
    print(f"Cargando archivo Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # Configuración de hojas de origen
    origin_sheets = {
        'peninsula': 'Origen Península',
        'canarias': 'Origen Canarias',
        'baleares': 'Origen Baleares',
        'islas_portugal': 'Origen Islas Portugal'
    }

    # Estructura final
    products_data = {}

    # Procesar cada hoja de origen
    for origin_type, sheet_name in origin_sheets.items():
        print(f"\nProcesando: {sheet_name}")
        ws = wb[sheet_name]

        # Extraer productos de esta hoja
        products = extract_products_from_sheet(ws, origin_type)

        print(f"  Productos encontrados: {len(products)}")

        # Integrar en la estructura global
        for product_name, config in products.items():
            if product_name not in products_data:
                products_data[product_name] = {
                    'id': product_name.replace(' ', '_').replace('(', '').replace(')', '').upper(),
                    'name': product_name,
                    'origins': {}
                }

            products_data[product_name]['origins'][origin_type] = config

    # Convertir a lista de productos
    products_list = []
    for product_name, data in products_data.items():
        product = {
            'id': data['id'],
            'name': data['name'],
            'origins': []
        }

        # Convertir origins a la estructura esperada
        for origin_type, config in data['origins'].items():
            # Por ahora, crear una configuración simplificada
            # TODO: Mapear destinos a zonas geográficas reales
            origin_config = {
                'origin_type': origin_type,
                'mapConfig': []  # Requiere mapeo manual de destinos a zonas
            }
            product['origins'].append(origin_config)

        products_list.append(product)

    # Guardar JSON
    output_data = {
        'products': products_list
    }

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)

    print(f"\n✓ Archivo generado: {output_path}")
    print(f"  Total productos: {len(products_list)}")

    # Mostrar resumen
    print("\nResumen de productos:")
    for product in products_list[:10]:  # Mostrar solo los primeros 10
        print(f"  - {product['name']} ({len(product['origins'])} orígenes)")


if __name__ == '__main__':
    excel_file = 'Portfolio ES-PT- IBÉRICO 24.09.2025 (v12).xlsx'
    output_file = 'scripts/map-config-extracted.json'

    generate_map_config(excel_file, output_file)

    print("\n" + "=" * 80)
    print("NOTA IMPORTANTE:")
    print("El archivo generado es una estructura base.")
    print("Requiere mapeo manual de destinos a zonas geográficas específicas.")
    print("Revisa scripts/map-config-extracted.json para ver la estructura.")
    print("=" * 80)
